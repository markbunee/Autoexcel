import pandas as pd
import difflib
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import time

def update_file_comparison(file1_path, file2_path, output_path, 
                          name_similarity_threshold=80, text_similarity_threshold=0.4,
                          file1_drug_col=1, file1_company_col=4, file1_status_col=7, file1_content_col=23,
                          file2_drug_col=0, file2_company_col=3, file2_status_col=6,
                          output_callback=None):
    """一致性评价进度文件对比更新功能"""
    
    def _print(msg):
        if output_callback:
            output_callback(msg)
        else:
            print(msg)
    
    def normalize_text(text):
        """文本规范化处理"""
        return str(text).strip().replace(" ", "").replace("（", "(").replace("）", ")")

    def calculate_text_similarity(text1, text2):
        """计算两个文本的相似度（0-1）"""
        return difflib.SequenceMatcher(None, str(text1), str(text2)).ratio()

    # ========== 参数配置 ==========
    NAME_SIM_THRESHOLD = name_similarity_threshold  # 名称相似度阈值
    TEXT_SIM_THRESHOLD = text_similarity_threshold  # 文本相似度阈值

    # 列索引设置
    FILE1_DRUG_COL = file1_drug_col
    FILE1_COMPANY_COL = file1_company_col
    FILE1_STATUS_COL = file1_status_col
    FILE1_CONTENT_COL = file1_content_col

    FILE2_DRUG_COL = file2_drug_col
    FILE2_COMPANY_COL = file2_company_col
    FILE2_STATUS_COL = file2_status_col

    # ========== 1. 数据准备 ==========
    _print("📂 读取数据文件中...")
    start_time = time.time()

    df1 = pd.read_excel(file1_path, dtype=str).fillna("")
    df2 = pd.read_excel(file2_path, dtype=str).fillna("")

    # 创建规范化的身份标识
    df1["标识"] = df1.apply(lambda row: f"{normalize_text(row.iloc[FILE1_DRUG_COL])}|{normalize_text(row.iloc[FILE1_COMPANY_COL])}", axis=1)
    df2["标识"] = df2.apply(lambda row: f"{normalize_text(row.iloc[FILE2_DRUG_COL])}|{normalize_text(row.iloc[FILE2_COMPANY_COL])}", axis=1)

    # ========== 2. 身份标识匹配 ==========
    _print("🔍 正在进行身份标识匹配...")
    matched_data = []
    unmatched_data = []

    # 构建file1的标识映射
    file1_ids = {row["标识"]: idx for idx, row in df1.iterrows()}

    for idx2, row2 in df2.iterrows():
        best_match = None
        highest_score = 0
        
        # 优先尝试精确匹配
        if row2["标识"] in file1_ids:
            best_match = file1_ids[row2["标识"]]
            highest_score = 100
        else:
            # 模糊匹配
            for id1, idx1 in file1_ids.items():
                # 拆分药物名称和企业名称
                drug2, comp2 = row2["标识"].split("|")
                drug1, comp1 = id1.split("|")
                
                # 分别计算相似度
                drug_sim = fuzz.ratio(drug1, drug2)
                comp_sim = fuzz.ratio(comp1, comp2)
                avg_sim = (drug_sim + comp_sim) / 2
                
                if avg_sim > highest_score and avg_sim >= NAME_SIM_THRESHOLD:
                    highest_score = avg_sim
                    best_match = idx1
        
        # 处理匹配结果
        if best_match is not None:
            row1 = df1.iloc[best_match]
            
            # 检查评审状态变化
            status1 = row1.iloc[FILE1_STATUS_COL]
            status2 = row2.iloc[FILE2_STATUS_COL]
            
            status_sim = calculate_text_similarity(status1, status2)
            status_changed = status_sim < TEXT_SIM_THRESHOLD
            
            # 获取文件一的第四列内容用于标准备注
            content1 = row1.iloc[FILE1_CONTENT_COL]
            
            # 构建匹配结果
            match_result = {
                "file1_index": best_match,
                "file2_index": idx2,
                "match_score": highest_score,
                "status_similarity": status_sim,
                "status_changed": status_changed,
                "file1_drug": row1.iloc[FILE1_DRUG_COL],
                "file1_company": row1.iloc[FILE1_COMPANY_COL],
                "file1_status": status1,
                "file2_drug": row2.iloc[FILE2_DRUG_COL],
                "file2_company": row2.iloc[FILE2_COMPANY_COL],
                "file2_status": status2,
                "file1_content": content1  # 新增：文件一的第四列内容
            }
            matched_data.append(match_result)
        else:
            unmatched_data.append({
                "file2_index": idx2,
                "drug": row2.iloc[FILE2_DRUG_COL],
                "company": row2.iloc[FILE2_COMPANY_COL],
                "status": row2.iloc[FILE2_STATUS_COL]
            })

    # ========== 3. 结果标记 ==========
    _print("🖌️ 标记结果文件中...")
    wb = load_workbook(file2_path)
    ws = wb.active

    # 创建样式
    new_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 新增-浅蓝
    changed_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 状态变更-黄色
    name_changed_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 名称变更-浅绿
    both_changed_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 名称+状态变更-橙色

    # 添加两列新列：匹配状态和标准备注
    max_col = ws.max_column + 1
    ws.cell(row=1, column=max_col, value="匹配状态")
    max_col += 1
    ws.cell(row=1, column=max_col, value="标准备注")

    # 初始化统计计数器
    exact_match_count = 0
    name_change_count = 0
    status_change_count = 0
    both_change_count = 0

    # 标记匹配结果
    for match in matched_data:
        row_idx = match["file2_index"] + 2  # Excel行号（标题行+1）
        remark_content = ""  # 标准备注列的内容
        row_color = None  # 行的颜色
        
        if match["match_score"] == 100:
            exact_match_count += 1
            if match["status_changed"]:
                remark = f"状态变更: {match['file1_status']} → {match['file2_status']}"
                # 整行标记黄色
                row_color = changed_fill
                remark_content = "2025-07-23至2025-08-04状态变化"  # 标准备注内容
                status_change_count += 1
            else:
                remark = "匹配成功"
                # 没有颜色标记
                remark_content = match["file1_content"]  # 使用文件一的第四列内容
        else:
            name_change_count += 1
            drug_change = f"{match['file1_drug']}→{match['file2_drug']}" if match["file1_drug"] != match["file2_drug"] else ""
            comp_change = f"{match['file1_company']}→{match['file2_company']}" if match["file1_company"] != match["file2_company"] else ""
            
            changes = [c for c in [drug_change, comp_change] if c]
            change_text = ", ".join(changes)
            
            if match["status_changed"]:
                both_change_count += 1
                remark = f"名称变更({change_text}) + 状态变更({match['file1_status']}→{match['file2_status']})"
                # 整行标记橙色
                row_color = both_changed_fill
                # 标准备注内容使用匹配状态列的内容
                remark_content = f"名称变更({change_text}) + 状态变更({match['file1_status']}→{match['file2_status']})"
            else:
                remark = f"名称变更({change_text})"
                # 整行标记浅绿色
                row_color = name_changed_fill
                # 标准备注内容使用匹配状态列的内容
                remark_content = f"名称变更({change_text})"
        
        # 写入匹配状态列
        ws.cell(row=row_idx, column=max_col-1, value=remark)
        
        # 写入标准备注列
        ws.cell(row=row_idx, column=max_col, value=remark_content)
        
        # 应用行颜色（如果有）
        if row_color:
            for col in range(1, max_col-1):  # 只对原有列应用颜色，不包括新添加的两列
                ws.cell(row=row_idx, column=col).fill = row_color

    # 标记未匹配数据（新增记录）
    for unmatched in unmatched_data:
        row_idx = unmatched["file2_index"] + 2
        remark = "新增记录"
        
        # 写入匹配状态列
        ws.cell(row=row_idx, column=max_col-1, value=remark)
        
        # 写入标准备注列
        remark_content = "2025-07-23后新增记录"
        ws.cell(row=row_idx, column=max_col, value=remark_content)
        
        # 整行标记浅蓝色（不包括新添加的两列）
        for col in range(1, max_col-1):
            ws.cell(row=row_idx, column=col).fill = new_fill

    # ========== 4. 保存结果 ==========
    wb.save(output_path)
    elapsed_time = time.time() - start_time
    _print(f"✅ 处理完成! 耗时: {elapsed_time:.2f}秒")
    _print(f"结果已保存至: {os.path.abspath(output_path)}")

    # ========== 5. 控制台输出详细统计信息 ==========
    _print("\n📊 详细匹配统计:")
    _print(f"文件1总记录数: {len(df1)}")
    _print(f"文件2总记录数: {len(df2)}")
    _print(f"成功匹配记录: {len(matched_data)}")
    _print(f"  ├─ 名称完全匹配: {exact_match_count}")
    _print(f"  │    ├─ 状态未变更: {exact_match_count - status_change_count}")
    _print(f"  │    └─ 状态变更: {status_change_count}")
    _print(f"  └─ 名称变更: {name_change_count}")
    _print(f"        ├─ 仅名称变更: {name_change_count - both_change_count}")
    _print(f"        └─ 名称+状态同时变更: {both_change_count}")
    _print(f"新增记录: {len(unmatched_data)}")

    # 验证统计一致性
    total_matched = exact_match_count + name_change_count
    if total_matched != len(matched_data):
        _print(f"⚠️ 警告: 匹配记录统计不一致 (计算:{total_matched} vs 实际:{len(matched_data)})")
    else:
        _print("✓ 匹配记录统计一致")

    total_records = len(matched_data) + len(unmatched_data)
    if total_records != len(df2):
        _print(f"⚠️ 警告: 总记录数不一致 (计算:{total_records} vs 文件2:{len(df2)})")
    else:
        _print("✓ 总记录数统计一致")
        
    return True