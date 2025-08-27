import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def clean_excel_data(source_file, output_file=None, clean_symbols=False, symbols_to_remove=None, 
                     mark_empty=True, mark_duplicates=True, clean_internal_spaces=False, 
                     clean_chinese_space=False, clean_english_punctuation=False, output_callback=None):
    """
    对Excel文件进行数据清洗
    
    Args:
        source_file (str): 源Excel文件路径
        output_file (str): 输出文件路径，默认为在原文件名后添加"_cleaned"
        clean_symbols (bool): 是否清除单元格中的多余符号，默认False
        symbols_to_remove (list): 要清除的符号列表，默认None表示清除空格、回车等
        mark_empty (bool): 是否标记空值单元格为黄色，默认True
        mark_duplicates (bool): 是否标记重复行整行蓝色，默认True
        clean_internal_spaces (bool): 是否清除单元格内部的空格和回车，默认False
        clean_chinese_space (bool): 是否清除中文全角空格，默认False
        clean_english_punctuation (bool): 是否处理英文标点符号差异，默认False
        output_callback (function): 输出回调函数，用于将日志信息传递给GUI
        
    Returns:
        bool: 清洗是否成功
    """
    def _print(msg):
        """内部打印函数，支持GUI输出"""
        if output_callback:
            output_callback(msg)
        else:
            print(msg)
    
    try:
        # 检查源文件是否存在
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"源文件不存在: {source_file}")
        
        # 设置输出文件名
        if output_file is None:
            name, ext = os.path.splitext(source_file)
            output_file = f"{name}_cleaned{ext}"
        
        # 读取Excel文件
        df = pd.read_excel(source_file, dtype=str)  # 使用字符串类型避免类型转换问题
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cleaned Data"
        
        # 定义样式
        empty_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
        duplicate_fill = PatternFill(start_color="e3fdfd", end_color="e3fdfd", fill_type="solid")  # 蓝色
        
        # 清除多余符号
        if clean_symbols:
            df = _clean_symbols_in_dataframe(df, symbols_to_remove)
        
        # 清除单元格内部空格和回车
        if clean_internal_spaces:
            df = _clean_internal_spaces_in_dataframe(df, _print, clean_chinese_space, clean_english_punctuation)
        
        # 将数据写入工作表
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 标记空值单元格
        if mark_empty:
            _mark_empty_cells(ws, empty_fill)
        
        # 标记重复行
        if mark_duplicates:
            _mark_duplicate_rows(ws, df, duplicate_fill)
        
        # 保存文件
        wb.save(output_file)
        _print(f"数据清洗完成，结果保存在: {output_file}")
        return True
        
    except Exception as e:
        _print(f"数据清洗时出错: {str(e)}")
        return False

def _clean_symbols_in_dataframe(df, symbols_to_remove):
    """
    清除DataFrame中指定的符号
    
    Args:
        df (DataFrame): 要清洗的DataFrame
        symbols_to_remove (list): 要清除的符号列表
        
    Returns:
        DataFrame: 清洗后的DataFrame
    """
    # 默认清除空格和回车
    if symbols_to_remove is None:
        symbols_to_remove = [' ', '\n', '\r']
    
    # 对每个单元格进行清洗
    for col in df.columns:
        for symbol in symbols_to_remove:
            df[col] = df[col].astype(str).str.replace(symbol, '', regex=False)
        
        # 处理NaN值
        df[col] = df[col].replace('nan', '')
    
    return df

def _clean_internal_spaces_in_dataframe(df, print_func, clean_chinese_space=False, clean_english_punctuation=False):
    """
    清除DataFrame中单元格内部的空格和回车
    
    Args:
        df (DataFrame): 要清洗的DataFrame
        print_func (function): 打印函数
        clean_chinese_space (bool): 是否清除中文全角空格
        clean_english_punctuation (bool): 是否处理英文标点符号差异
        
    Returns:
        DataFrame: 清洗后的DataFrame
    """
    print_func("开始清除单元格内部空格和回车...")
    
    # 对每个单元格进行清洗，将多个连续的空格和回车替换为单个空格，并去除首尾空格
    for col in df.columns:
        # 将换行符、回车符替换为空格
        df[col] = df[col].astype(str).str.replace('[\n\r]+', ' ', regex=True)
        
        # 根据选项决定是否处理中文全角空格
        if clean_chinese_space:
            # 替换各种类型的空格（包括中文全角空格）为标准空格
            df[col] = df[col].astype(str).str.replace('[\s　]+', ' ', regex=True)
        else:
            # 只处理标准空格和制表符
            df[col] = df[col].astype(str).str.replace('[ \t]+', ' ', regex=True)
        
        # 去除首尾空格
        df[col] = df[col].astype(str).str.strip()
        
        # 处理NaN值
        df[col] = df[col].replace('nan', '')
    
    print_func("单元格内部空格和回车清除完成")
    return df

def _mark_empty_cells(worksheet, fill_style):
    """
    标记空值单元格为黄色
    
    Args:
        worksheet: Excel工作表对象
        fill_style: 填充样式
    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == '':
                cell.fill = fill_style

def _mark_duplicate_rows(worksheet, df, fill_style):
    """
    标记重复行整行蓝色
    
    Args:
        worksheet: Excel工作表对象
        df (DataFrame): 原始数据DataFrame
        fill_style: 填充样式
    """
    # 找到重复行（keep=False表示标记所有重复项，包括第一次出现的）
    duplicate_rows = df.duplicated(keep=False)
    
    # 标记重复行
    header_row = 1  # 假设第一行是标题行
    for idx, is_duplicate in enumerate(duplicate_rows, start=header_row + 1):
        if is_duplicate:
            # 标记整行
            for cell in worksheet[idx]:
                cell.fill = fill_style

def clean_excel_advanced(source_file, output_file=None, clean_configs=None, output_callback=None):
    """
    高级数据清洗功能，支持多种清洗配置
    
    Args:
        source_file (str): 源Excel文件路径
        output_file (str): 输出文件路径
        clean_configs (dict): 清洗配置字典，支持以下键:
            - 'empty_cells': 标记空值单元格配置
            - 'duplicate_rows': 标记重复行配置
            - 'symbols': 符号清洗配置
        output_callback (function): 输出回调函数，用于将日志信息传递给GUI
            
    Returns:
        bool: 清洗是否成功
    """
    def _print(msg):
        """内部打印函数，支持GUI输出"""
        if output_callback:
            output_callback(msg)
        else:
            print(msg)
    
    try:
        # 检查源文件是否存在
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"源文件不存在: {source_file}")
        
        # 设置输出文件名
        if output_file is None:
            name, ext = os.path.splitext(source_file)
            output_file = f"{name}_cleaned{ext}"
        
        # 读取Excel文件
        df = pd.read_excel(source_file, dtype=str)
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cleaned Data"
        
        # 默认配置
        if clean_configs is None:
            clean_configs = {
                'empty_cells': {'enabled': True, 'color': 'FFFF00'},
                'duplicate_rows': {'enabled': True, 'color': 'e3fdfd'},
                'symbols': {'enabled': False, 'symbols': [' ', '\n', '\r']}
            }
        
        # 处理符号清洗
        if clean_configs.get('symbols', {}).get('enabled', False):
            symbols = clean_configs['symbols'].get('symbols', [' ', '\n', '\r'])
            df = _clean_symbols_in_dataframe(df, symbols)
        
        # 将数据写入工作表
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 处理空值单元格标记
        if clean_configs.get('empty_cells', {}).get('enabled', True):
            color = clean_configs['empty_cells'].get('color', 'FFFF00')
            fill_style = PatternFill(start_color=color, end_color=color, fill_type="solid")
            _mark_empty_cells(ws, fill_style)
        
        # 处理重复行标记
        if clean_configs.get('duplicate_rows', {}).get('enabled', True):
            color = clean_configs['duplicate_rows'].get('color', 'e3fdfd')
            fill_style = PatternFill(start_color=color, end_color=color, fill_type="solid")
            _mark_duplicate_rows(ws, df, fill_style)
        
        # 保存文件
        wb.save(output_file)
        _print(f"高级数据清洗完成，结果保存在: {output_file}")
        return True
        
    except Exception as e:
        _print(f"数据清洗时出错: {str(e)}")
        return False

# 示例用法
if __name__ == "__main__":
    # 示例1: 基本数据清洗
    # clean_excel_data(
    #     source_file="data.xlsx",
    #     clean_symbols=True,
    #     symbols_to_remove=['*', '#', '@'],
    #     mark_empty=True,
    #     mark_duplicates=True
    # )
    
    # 示例2: 高级数据清洗
    # clean_configs = {
    #     'empty_cells': {'enabled': True, 'color': 'FFFF00'},  # 黄色标记空值
    #     'duplicate_rows': {'enabled': True, 'color': 'e3fdfd'},  # 蓝色标记重复行
    #     'symbols': {'enabled': True, 'symbols': ['*', ' ', '\n']}  # 清除指定符号
    # }
    # clean_excel_advanced("data.xlsx", "cleaned_data.xlsx", clean_configs)
    pass